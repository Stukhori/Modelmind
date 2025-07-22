from flask import Flask, render_template, request, redirect, flash, url_for, session
import pandas as pd
import os
import re
from werkzeug.utils import secure_filename
import google.generativeai as genai
import logging
import openpyxl
from uuid import uuid4
from datetime import datetime
import markdown
import mimetypes
mimetypes.init()


logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


app = Flask(__name__)
app.secret_key = os.urandom(24)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['ALLOWED_EXTENSIONS'] = {'xls', 'xlsx'}
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
app.config['SESSION_PERMANENT'] = False


genai.configure(api_key="AIzaSyCxnBvgM-EJyQblgahPSKGTvr5iI-sb51c")


if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])


def allowed_file(filename):
    # Check file extension
    ext_ok = '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']
   
    # Check MIME type
    mime_type = mimetypes.guess_type(filename)[0]
    mime_ok = mime_type in [
        'application/vnd.ms-excel',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    ] if mime_type else False
   
    return ext_ok and mime_ok


def detect_errors(df: pd.DataFrame, file_path: str, sheet_name: str) -> str:
    error_report = ""
    if df.isnull().values.any():
        error_report += f"Missing values detected in {df.isnull().sum().to_string()}.\n"
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('#'):
                    error_report += f"Formula error in cell {cell.coordinate}: {cell.value}.\n"
    except Exception as e:
        logger.error(f"Error detecting Excel errors: {e}")
    return error_report.strip() or "No obvious errors detected."


def compute_trends(df: pd.DataFrame) -> str:
    trend_summary = ""
    numeric_cols = df.select_dtypes(include='number').columns
    if 'Date' in df.columns or any(col.lower().find('date') >= 0 for col in df.columns):
        date_col = next((col for col in df.columns if 'date' in col.lower() or col == 'Date'), None)
        if date_col:
            try:
                df[date_col] = pd.to_datetime(df[date_col])
                for col in numeric_cols:
                    if col != date_col:
                        yearly = df.groupby(df[date_col].dt.year)[col].sum()
                        if len(yearly) > 1:
                            growth = yearly.pct_change().mean() * 100
                            trend_summary += f"{col} has {'increased' if growth > 0 else 'decreased'} on average by {abs(growth):.2f}% yearly.\n"
            except Exception as e:
                logger.error(f"Error computing trends: {e}")
    return trend_summary.strip() or "No clear time-series trends detected."


def get_formula(file_path: str, sheet_name: str, cell_ref: str) -> str:
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]
        cell = ws[cell_ref]
        return cell.formula if cell.formula else f"No formula in cell {cell_ref}."
    except Exception as e:
        logger.error(f"Error retrieving formula: {e}")
        return f"Error retrieving formula for cell {cell_ref}."


def summarize_excel(df: pd.DataFrame, sheet_name: str = None) -> str:
    try:
        summary = ""
        if sheet_name:
            summary += f"Sheet: {sheet_name}\n"
           
        if df.shape[0] > 10 or df.shape[1] > 10:
            summary += f"The spreadsheet has {df.shape[0]} rows and {df.shape[1]} columns. "
            summary += f"Columns: {', '.join(df.columns)}. "
            summary += f"Summary statistics:\n{df.describe().to_string()}"
        else:
            summary += f"The spreadsheet has the following data:\n{df.head().to_string()}"
        return summary
    except Exception as e:
        logger.error(f"Error summarizing Excel: {e}")
        return "Error summarizing Excel data."


def analyze_multiple_sheets(file_path: str, sheet_names: list) -> dict:
    analysis_results = {
        'combined_summary': '',
        'individual_sheets': {},
        'cross_sheet_analysis': ''
    }
   
    try:
        logger.info(f"Analyzing {len(sheet_names)} sheets: {', '.join(sheet_names)}")
       
        all_sheets = pd.read_excel(file_path, sheet_name=sheet_names, engine='openpyxl')
       
        for sheet in sheet_names:
            logger.info(f"Processing sheet: {sheet}")
            df = all_sheets[sheet]
            analysis_results['individual_sheets'][sheet] = {
                'summary': summarize_excel(df, sheet),
                'errors': detect_errors(df, file_path, sheet),
                'trends': compute_trends(df),
                'shape': df.shape,
                'numeric_cols': df.select_dtypes(include='number').columns.tolist()
            }
       
        analysis_results['combined_summary'] = f"File contains {len(sheet_names)} sheets: {', '.join(sheet_names)}\n"
       
        if len(sheet_names) > 1:
            common_columns = set(all_sheets[sheet_names[0]].columns)
            for sheet in sheet_names[1:]:
                common_columns.intersection_update(all_sheets[sheet].columns)
           
            if common_columns:
                analysis_results['cross_sheet_analysis'] = f"Common columns across sheets: {', '.join(common_columns)}\n"
           
                numeric_comparison = ""
                for col in common_columns:
                    if all_sheets[sheet_names[0]][col].dtype in ['int64', 'float64']:
                        col_values = {}
                        for sheet in sheet_names:
                            col_values[sheet] = all_sheets[sheet][col].sum()
                       
                        if len(set(col_values.values())) > 1:
                            max_sheet = max(col_values.items(), key=lambda x: x[1])
                            min_sheet = min(col_values.items(), key=lambda x: x[1])
                            numeric_comparison += (
                                f"Column '{col}' totals vary across sheets. "
                                f"Highest in {max_sheet[0]} ({max_sheet[1]:,.2f}), "
                                f"lowest in {min_sheet[0]} ({min_sheet[1]:,.2f})\n"
                            )
               
                if numeric_comparison:
                    analysis_results['cross_sheet_analysis'] += numeric_comparison
            else:
                analysis_results['cross_sheet_analysis'] = "No common columns found across sheets."
        else:
            analysis_results['cross_sheet_analysis'] = "Only one sheet in the file."
       
        logger.info("Multi-sheet analysis completed")
        return analysis_results
   
    except Exception as e:
        logger.error(f"Error in multi-sheet analysis: {e}")
        analysis_results['error'] = f"Error analyzing multiple sheets: {str(e)}"
        return analysis_results


def call_gemini_api(prompt: str) -> str:
    try:
        logger.info("Calling Gemini API...")
        model = genai.GenerativeModel("gemini-1.5-flash")
        formatted_prompt = (
            f"{prompt}\n\n"
            "Please format your response with:\n"
            "- Clear section headings\n"
            "- Bullet points for lists\n"
            "- Paragraphs for detailed explanations\n"
            "- Bold text for key insights\n"
            "- Proper line breaks between sections\n"
            "Use Markdown formatting for best readability."
        )
        response = model.generate_content(formatted_prompt)
        logger.info("Gemini API call successful")
       
        # Convert Markdown to HTML
        html_content = markdown.markdown(response.text.strip())
        return html_content
    except Exception as e:
        logger.error(f"Error calling Gemini API: {e}")
        return markdown.markdown(f"Error generating Gemini response: {e}")


@app.route('/')
def home():
    session.pop('file_data', None)
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'excel_file' not in request.files or 'user_question' not in request.form:
        flash('Please provide both a file and a question.')
        return redirect(url_for('home'))


    file = request.files['excel_file']
    user_question = request.form['user_question'].strip()


    if not user_question:
        flash('Please provide a question.')
        return redirect(url_for('home'))


    if file.filename == '':
        flash('No file selected.')
        return redirect(url_for('home'))


    # Enhanced file validation
    if not allowed_file(file.filename):
        # Get detected MIME type for better error message
        mime_type = mimetypes.guess_type(file.filename)[0] or "unknown"
        flash(f'Invalid file format. Detected type: {mime_type}. Only Excel files (.xls, .xlsx) are allowed.')
        return redirect(url_for('home'))


    filename = secure_filename(file.filename)
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{uuid4()}_{filename}")
    file.save(file_path)


    try:
        logger.info(f"Processing file: {filename}")
        excel_file = pd.ExcelFile(file_path, engine='openpyxl')
        sheet_names = excel_file.sheet_names
        if not sheet_names:
            raise ValueError("Excel file has no sheets.")
       
        logger.info(f"File contains {len(sheet_names)} sheets: {', '.join(sheet_names)}")
           
        # Check if question mentions a specific sheet
        specific_sheet = None
        for sheet in sheet_names:
            if re.search(rf'\b{re.escape(sheet)}\b', user_question, re.IGNORECASE):
                specific_sheet = sheet
                break
           
        multi_sheet_keywords = ['across sheets', 'multiple sheets', 'all sheets', 'between sheets', 'each sheet']
        analyze_all_sheets = any(keyword in user_question.lower() for keyword in multi_sheet_keywords)
       
        # FIX: Handle specific sheet requests properly
        if specific_sheet:
            logger.info(f"User requested specific sheet: {specific_sheet}")
            df = pd.read_excel(file_path, engine='openpyxl', sheet_name=specific_sheet)
           
            num_rows = len(df)
            num_cols = len(df.columns)
            table_html = df.head(5).to_html(index=False, classes='data-table')
            numeric_sums = df.select_dtypes(include='number').sum().to_dict() if not df.empty else {}
           
            error_report = detect_errors(df, file_path, specific_sheet)
            trend_summary = compute_trends(df)
           
            data_summary = summarize_excel(df, specific_sheet)
            cell_ref = None
            if re.search(r'formula\s+in\s+cell\s+(\w+\d+)', user_question, re.IGNORECASE):
                cell_match = re.search(r'(\w+\d+)', user_question, re.IGNORECASE)
                if cell_match:
                    cell_ref = cell_match.group(1)
                    formula = get_formula(file_path, specific_sheet, cell_ref)
                    data_summary += f"\nFormula in cell {cell_ref}: {formula}"
           
            prompt = f"Data from Excel file (sheet: {specific_sheet}): {data_summary}\n"
            if error_report != "No obvious errors detected.":
                prompt += f"Errors detected: {error_report}\n"
            if trend_summary != "No clear time-series trends detected.":
                prompt += f"Trends: {trend_summary}\n"
            prompt += f"User question: {user_question}\nAnswer based on the provided data."
           
            logger.info(f"Gemini prompt:\n{prompt}")
            answer = call_gemini_api(prompt)
           
            session['file_data'] = {
                'file_path': file_path,
                'sheet_name': specific_sheet,
                'sheet_names': sheet_names,
                'data_summary': data_summary,
                'error_report': error_report,
                'trend_summary': trend_summary,
                'num_rows': num_rows,
                'num_cols': num_cols,
                'table_html': table_html,
                'numeric_sums': numeric_sums,
                'filename': filename,
                'current_sheet': specific_sheet,
                'multi_sheet': False
            }
            session['qa_history'] = [{'question': user_question, 'answer': answer}]
           
            return render_template('result.html',
                                filename=filename,
                                question=user_question,
                                answer=answer,
                                sheet_names=sheet_names,
                                num_rows=num_rows,
                                num_cols=num_cols,
                                numeric_sums=numeric_sums,
                                table_html=table_html,
                                current_sheet=specific_sheet,
                                trend_summary=trend_summary,  # Add this line
                                qa_history=session['qa_history'])
           
        elif len(sheet_names) > 1 and analyze_all_sheets:
            logger.info("Performing multi-sheet analysis")
            analysis_results = analyze_multiple_sheets(file_path, sheet_names)
           
            tables_html = {}
            for sheet in sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet, engine='openpyxl')
                tables_html[sheet] = df.head(5).to_html(index=False, classes='data-table')
           
            prompt = f"Excel file with multiple sheets analysis:\n"
            prompt += analysis_results['combined_summary'] + "\n"
            prompt += analysis_results['cross_sheet_analysis'] + "\n"
           
            for sheet in sheet_names:
                prompt += f"\nSheet '{sheet}' analysis:\n"
                prompt += f"- Summary: {analysis_results['individual_sheets'][sheet]['summary']}\n"
                if analysis_results['individual_sheets'][sheet]['errors'] != "No obvious errors detected.":
                    prompt += f"- Errors: {analysis_results['individual_sheets'][sheet]['errors']}\n"
                if analysis_results['individual_sheets'][sheet]['trends'] != "No clear time-series trends detected.":
                    prompt += f"- Trends: {analysis_results['individual_sheets'][sheet]['trends']}\n"
           
            prompt += f"\nUser question: {user_question}\nPlease provide a comprehensive answer considering all sheets."
           
            logger.info(f"Gemini prompt:\n{prompt}")
            answer = call_gemini_api(prompt)
           
            session['file_data'] = {
                'file_path': file_path,
                'sheet_names': sheet_names,
                'analysis_results': analysis_results,
                'tables_html': tables_html,
                'filename': filename,
                'current_sheet': None,
                'multi_sheet': True
            }
            session['qa_history'] = [{'question': user_question, 'answer': answer}]
           
            return render_template('result_multi.html',
                                filename=filename,
                                question=user_question,
                                answer=answer,
                                sheet_names=sheet_names,
                                analysis_results=analysis_results,
                                tables_html=tables_html,
                                qa_history=session['qa_history'])
        else:
            # Default to first sheet
            logger.info("Performing single sheet analysis")
            df = pd.read_excel(file_path, engine='openpyxl', sheet_name=sheet_names[0])
           
            num_rows = len(df)
            num_cols = len(df.columns)
            table_html = df.head(5).to_html(index=False, classes='data-table')
            numeric_sums = df.select_dtypes(include='number').sum().to_dict() if not df.empty else {}
           
            error_report = detect_errors(df, file_path, sheet_names[0])
            trend_summary = compute_trends(df)
           
            data_summary = summarize_excel(df, sheet_names[0])
            cell_ref = None
            if re.search(r'formula\s+in\s+cell\s+(\w+\d+)', user_question, re.IGNORECASE):
                cell_match = re.search(r'(\w+\d+)', user_question, re.IGNORECASE)
                if cell_match:
                    cell_ref = cell_match.group(1)
                    formula = get_formula(file_path, sheet_names[0], cell_ref)
                    data_summary += f"\nFormula in cell {cell_ref}: {formula}"
           
            prompt = f"Data from Excel file (sheet: {sheet_names[0]}): {data_summary}\n"
            if error_report != "No obvious errors detected.":
                prompt += f"Errors detected: {error_report}\n"
            if trend_summary != "No clear time-series trends detected.":
                prompt += f"Trends: {trend_summary}\n"
            prompt += f"User question: {user_question}\nAnswer based on the provided data."
           
            logger.info(f"Gemini prompt:\n{prompt}")
            answer = call_gemini_api(prompt)
           
            session['file_data'] = {
                'file_path': file_path,
                'sheet_name': sheet_names[0],
                'sheet_names': sheet_names,
                'data_summary': data_summary,
                'error_report': error_report,
                'trend_summary': trend_summary,
                'num_rows': num_rows,
                'num_cols': num_cols,
                'table_html': table_html,
                'numeric_sums': numeric_sums,
                'filename': filename,
                'current_sheet': sheet_names[0],
                'multi_sheet': False
            }
            session['qa_history'] = [{'question': user_question, 'answer': answer}]
           
            return render_template('result.html',
                                filename=filename,
                                question=user_question,
                                answer=answer,
                                sheet_names=sheet_names,
                                num_rows=num_rows,
                                num_cols=num_cols,
                                numeric_sums=numeric_sums,
                                table_html=table_html,
                                current_sheet=sheet_names[0],
                                qa_history=session['qa_history'])


    except Exception as e:
        logger.error(f"Error processing file: {str(e)}")
        if os.path.exists(file_path):
            os.remove(file_path)
        flash(f'Error processing file: {str(e)}')
        return redirect(url_for('home'))


@app.route('/switch_sheet', methods=['POST'])
def switch_sheet():
    if 'file_data' not in session:
        flash('No file data available. Please upload a file.')
        return redirect(url_for('home'))
       
    selected_sheet = request.form.get('sheet_name')
    if not selected_sheet:
        flash('No sheet selected.')
        return redirect(url_for('result'))
   
    file_data = session['file_data']
    if selected_sheet not in file_data['sheet_names']:
        flash('Invalid sheet selection.')
        return redirect(url_for('result'))
   
    try:
        logger.info(f"Switching to sheet: {selected_sheet}")
        file_path = file_data['file_path']
        df = pd.read_excel(file_path, engine='openpyxl', sheet_name=selected_sheet)
       
        session['file_data'].update({
            'current_sheet': selected_sheet,
            'sheet_name': selected_sheet,
            'data_summary': summarize_excel(df, selected_sheet),
            'error_report': detect_errors(df, file_path, selected_sheet),
            'trend_summary': compute_trends(df),
            'num_rows': len(df),
            'num_cols': len(df.columns),
            'table_html': df.head(5).to_html(index=False, classes='data-table'),
            'numeric_sums': df.select_dtypes(include='number').sum().to_dict() if not df.empty else {},
            'multi_sheet': False
        })
       
        return render_template('result.html',
                            filename=file_data['filename'],
                            question=session['qa_history'][-1]['question'] if 'qa_history' in session else "Sheet switched",
                            answer=f"Now viewing sheet: {selected_sheet}",
                            sheet_names=file_data['sheet_names'],
                            num_rows=len(df),
                            num_cols=len(df.columns),
                            numeric_sums=df.select_dtypes(include='number').sum().to_dict() if not df.empty else {},
                            table_html=df.head(5).to_html(index=False, classes='data-table'),
                            current_sheet=selected_sheet,
                            trend_summary=compute_trends(df),  # Add this line
                            qa_history=session.get('qa_history', []))
   
    except Exception as e:
        logger.error(f"Error switching sheets: {str(e)}")
        flash(f'Error switching sheets: {str(e)}')
        return redirect(url_for('result'))
   


@app.route('/ask_another', methods=['POST'])
def ask_another():
    if 'file_data' not in session or 'user_question' not in request.form:
        flash('No file data available or question missing. Please upload a file.')
        return redirect(url_for('home'))


    user_question = request.form['user_question'].strip()
    if not user_question:
        flash('Please provide a question.')
        return redirect(url_for('result'))


    try:
        file_data = session['file_data']
        file_path = file_data['file_path']
       
        specific_sheet = None
        for sheet in file_data['sheet_names']:
            if re.search(rf'\b{re.escape(sheet)}\b', user_question, re.IGNORECASE):
                specific_sheet = sheet
                break
       
        if specific_sheet:
            logger.info(f"Question about specific sheet: {specific_sheet}")
            df = pd.read_excel(file_path, engine='openpyxl', sheet_name=specific_sheet)
            data_summary = summarize_excel(df, specific_sheet)
            error_report = detect_errors(df, file_path, specific_sheet)
            trend_summary = compute_trends(df)
           
            prompt = f"Data from Excel file (sheet: {specific_sheet}): {data_summary}\n"
            if error_report != "No obvious errors detected.":
                prompt += f"Errors detected: {error_report}\n"
            if trend_summary != "No clear time-series trends detected.":
                prompt += f"Trends: {trend_summary}\n"
            prompt += f"User question: {user_question}\nAnswer based on the provided data."
           
            logger.info(f"Gemini prompt:\n{prompt}")
            answer = call_gemini_api(prompt)
           
            session['file_data'].update({
                'current_sheet': specific_sheet,
                'sheet_name': specific_sheet,
                'data_summary': data_summary,
                'error_report': error_report,
                'trend_summary': trend_summary,
                'num_rows': len(df),
                'num_cols': len(df.columns),
                'table_html': df.head(5).to_html(index=False, classes='data-table'),
                'numeric_sums': df.select_dtypes(include='number').sum().to_dict() if not df.empty else {},
                'multi_sheet': False
            })
           
            # Add to QA history
            session['qa_history'].append({'question': user_question, 'answer': answer})
            session.modified = True
           
            return render_template('result.html',
                                filename=file_data['filename'],
                                question=user_question,
                                answer=answer,
                                sheet_names=file_data['sheet_names'],
                                num_rows=len(df),
                                num_cols=len(df.columns),
                                numeric_sums=df.select_dtypes(include='number').sum().to_dict() if not df.empty else {},
                                table_html=df.head(5).to_html(index=False, classes='data-table'),
                                current_sheet=specific_sheet,
                                qa_history=session['qa_history'])
        elif file_data.get('multi_sheet', False) or any(kw in user_question.lower() for kw in ['all sheets', 'across sheets', 'multiple sheets']):
            logger.info("Performing full multi-sheet analysis")
            analysis_results = analyze_multiple_sheets(file_path, file_data['sheet_names'])
           
            session['file_data']['analysis_results'] = analysis_results
           
            tables_html = {}
            for sheet in file_data['sheet_names']:
                df = pd.read_excel(file_path, sheet_name=sheet, engine='openpyxl')
                tables_html[sheet] = df.head(5).to_html(index=False, classes='data-table')
           
            session['file_data']['tables_html'] = tables_html
           
            prompt = f"Excel file with multiple sheets analysis:\n"
            prompt += analysis_results['combined_summary'] + "\n"
            prompt += analysis_results['cross_sheet_analysis'] + "\n"
           
            for sheet in file_data['sheet_names']:
                prompt += f"\nSheet '{sheet}' analysis:\n"
                prompt += f"- Summary: {analysis_results['individual_sheets'][sheet]['summary']}\n"
                if analysis_results['individual_sheets'][sheet]['errors'] != "No obvious errors detected.":
                    prompt += f"- Errors: {analysis_results['individual_sheets'][sheet]['errors']}\n"
                if analysis_results['individual_sheets'][sheet]['trends'] != "No clear time-series trends detected.":
                    prompt += f"- Trends: {analysis_results['individual_sheets'][sheet]['trends']}\n"
           
            prompt += f"\nUser question: {user_question}\nPlease provide a comprehensive answer considering all sheets."
           
            logger.info(f"Gemini prompt:\n{prompt}")
            answer = call_gemini_api(prompt)
           
            session['qa_history'].append({'question': user_question, 'answer': answer})
            session.modified = True
           
            return render_template('result_multi.html',
                                filename=file_data['filename'],
                                question=user_question,
                                answer=answer,
                                sheet_names=file_data['sheet_names'],
                                analysis_results=analysis_results,
                                tables_html=tables_html,
                                qa_history=session['qa_history'])
        else:
            logger.info("Performing single sheet analysis")
            current_sheet = file_data.get('current_sheet', file_data['sheet_names'][0])
           
            df = pd.read_excel(file_path, engine='openpyxl', sheet_name=current_sheet)
           
            data_summary = summarize_excel(df, current_sheet)
            error_report = detect_errors(df, file_path, current_sheet)
            trend_summary = compute_trends(df)
           
            cell_ref = None
            if re.search(r'formula\s+in\s+cell\s+(\w+\d+)', user_question, re.IGNORECASE):
                cell_match = re.search(r'(\w+\d+)', user_question, re.IGNORECASE)
                if cell_match:
                    cell_ref = cell_match.group(1)
                    formula = get_formula(file_path, current_sheet, cell_ref)
                    data_summary += f"\nFormula in cell {cell_ref}: {formula}"
           
            prompt = f"Data from Excel file (sheet: {current_sheet}): {data_summary}\n"
            if error_report != "No obvious errors detected.":
                prompt += f"Errors detected: {error_report}\n"
            if trend_summary != "No clear time-series trends detected.":
                prompt += f"Trends: {trend_summary}\n"
            prompt += f"User question: {user_question}\nAnswer based on the provided data."
           
            logger.info(f"Gemini prompt:\n{prompt}")
            answer = call_gemini_api(prompt)
           
            session['file_data'].update({
                'data_summary': data_summary,
                'error_report': error_report,
                'trend_summary': trend_summary,
                'num_rows': len(df),
                'num_cols': len(df.columns),
                'table_html': df.head(5).to_html(index=False, classes='data-table'),
                'numeric_sums': df.select_dtypes(include='number').sum().to_dict() if not df.empty else {}
            })
           
            session['qa_history'].append({'question': user_question, 'answer': answer})
            session.modified = True
           
            return render_template('result.html',
                                filename=file_data['filename'],
                                question=user_question,
                                answer=answer,
                                sheet_names=file_data['sheet_names'],
                                num_rows=len(df),
                                num_cols=len(df.columns),
                                numeric_sums=df.select_dtypes(include='number').sum().to_dict() if not df.empty else {},
                                table_html=df.head(5).to_html(index=False, classes='data-table'),
                                current_sheet=current_sheet,
                                trend_summary=trend_summary,  # Add this line
                                qa_history=session['qa_history'])


    except Exception as e:
        logger.error(f"Error processing question: {str(e)}")
        flash(f'Error processing question: {str(e)}')
        return redirect(url_for('home'))


@app.route('/result')
def result():
    if 'file_data' not in session or 'qa_history' not in session:
        flash('No session data available. Please upload a file.')
        return redirect(url_for('home'))
       
    file_data = session['file_data']
    qa_history = session['qa_history']
   
    if file_data.get('multi_sheet', False):
        return render_template('result_multi.html',
                            filename=file_data['filename'],
                            question=qa_history[-1]['question'],
                            answer=qa_history[-1]['answer'],
                            sheet_names=file_data['sheet_names'],
                            analysis_results=file_data['analysis_results'],
                            tables_html=file_data['tables_html'],
                            qa_history=qa_history)
    else:
        return render_template('result.html',
                            filename=file_data['filename'],
                            question=qa_history[-1]['question'],
                            answer=qa_history[-1]['answer'],
                            sheet_names=file_data['sheet_names'],
                            num_rows=file_data['num_rows'],
                            num_cols=file_data['num_cols'],
                            numeric_sums=file_data['numeric_sums'],
                            table_html=file_data['table_html'],
                            current_sheet=file_data.get('current_sheet', file_data['sheet_names'][0]),
                            qa_history=qa_history)


if __name__ == '__main__':
    print("Starting Flask app...")
    app.run(host='0.0.0.0', port=8080, debug=True)